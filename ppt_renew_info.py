import win32com.client
import os
from openpyxl import load_workbook
import time


def update_ppt_from_excel(ppt_path, excel_path):
    try:
        # 检查文件存在性
        if not os.path.exists(ppt_path) or not os.path.exists(excel_path):
            print("文件不存在，终止操作")
            return

        # 读取Excel数据（保留content，供文本框使用）
        wb = load_workbook(excel_path, data_only=True)
        ws = wb["汇总"]
        data_index = {}
        print(f"\n===== 读取Excel数据 =====")
        for row in ws.iter_rows(min_row=2, values_only=True):
            # 幻灯片编号转换为整数
            try:
                slide_num = int(row[0]) if row[0] else None
            except:
                slide_num = None
            shape_type = str(row[1]).strip() if row[1] else ""
            shape_name = str(row[2]).strip() if row[2] else ""
            content = str(row[3]).strip() if row[3] else ""
            data_cols = row[4:8]

            if not slide_num or not shape_name:
                print(f"  跳过无效行：幻灯片编号{row[0]}，形状{shape_name}")
                continue

            # 清洗数据（保留content，同时处理data）
            clean_data = [str(d).strip() if d is not None else "" for d in data_cols]
            if slide_num not in data_index:
                data_index[slide_num] = {}
            if shape_name not in data_index[slide_num]:
                data_index[slide_num][shape_name] = []
            data_index[slide_num][shape_name].append({
                "type": shape_type,
                "content": content,  # 保留content，供文本框使用
                "data": clean_data
            })
            print(f"  加载数据：幻灯片{slide_num}，形状{shape_name}，类型{shape_type}")

        # 启动PPT
        ppt_app = win32com.client.Dispatch("PowerPoint.Application")
        ppt_app.Visible = True
        ppt_app.DisplayAlerts = False
        time.sleep(1)
        presentation = ppt_app.Presentations.Open(ppt_path)
        time.sleep(1)
        print(f"\n===== 开始更新PPT =====")

        # 遍历幻灯片
        for slide in presentation.Slides:
            slide_num = slide.SlideNumber
            print(f"\n----- 处理幻灯片 {slide_num} -----")

            if slide_num not in data_index:
                print("  无对应Excel数据，跳过")
                continue
            shape_data = data_index[slide_num]

            # 遍历形状
            for shape in slide.Shapes:
                shape_name = shape.Name
                if shape_name not in shape_data:
                    continue

                updates = shape_data[shape_name]
                print(f"  更新形状：{shape_name}（{len(updates)}条数据）")

                # 2. 表格更新（仅用data，忽略content）
                if shape.HasTable:
                    table = shape.Table
                    ppt_rows = table.Rows.Count
                    ppt_cols = table.Columns.Count
                    print(f"    表格结构：{ppt_rows}行×{ppt_cols}列")

                    for data_idx, update in enumerate(updates):
                        if "表格" in update["type"] and (data_idx + 1) <= ppt_rows:
                            row_idx = data_idx + 1  # PPT表格行索引从1开始
                            for col_idx in range(min(len(update["data"]), ppt_cols)):
                                table_col = col_idx + 1  # 列索引从1开始
                                table.Cell(row_idx, table_col).Shape.TextFrame.TextRange.Text = update["data"][col_idx]
                            print(f"    表格行{row_idx}更新完成：{update['data']}")

                # 1. 文本框更新（依赖content数据）
                elif shape.HasTextFrame and shape.TextFrame.HasText:
                    for update in updates:
                        if "文本框" in update["type"].replace(" ", ""):  # 兼容带空格的类型
                            shape.TextFrame.TextRange.Text = update["content"]
                            print(f"    文本框更新：{update['content'][:50]}...")

                # 3. 图表更新（仅用data，忽略content）
                elif shape.HasChart:
                    chart = shape.Chart
                    try:
                        # 激活数据源
                        chart_data = chart.ChartData
                        activate_success = False
                        for _ in range(3):
                            try:
                                chart_data.Activate()
                                time.sleep(0.5)
                                activate_success = True
                                break
                            except:
                                time.sleep(0.5)
                        if not activate_success:
                            print("    数据源激活失败，跳过")
                            continue

                        workbook = chart_data.Workbook
                        worksheet = workbook.Worksheets(1)
                        # 清除原有数据
                        try:
                            worksheet.UsedRange.ClearContents()
                            print("    已清除原有图表数据")
                        except:
                            print("    清除数据失败，直接覆盖")

                        # 筛选有效图表类型
                        valid_types = {"饼图", "柱形图", "条形图"}
                        chart_updates = [u for u in updates if u["type"] in valid_types]
                        if not chart_updates:
                            print("    无有效图表数据，跳过")
                            continue

                        # 写入图表数据
                        for row_idx, update in enumerate(chart_updates, 1):
                            for col_idx, data in enumerate(update["data"], 1):  # 从第1列开始写入
                                try:
                                    # 优先转换为数值
                                    worksheet.Cells(row_idx, col_idx).Value = float(data)
                                except:
                                    # 转换失败则按字符串写入
                                    worksheet.Cells(row_idx, col_idx).Value = data
                            print(f"    图表行{row_idx}更新完成：{update['data']}")

                        # 刷新图表并保存
                        chart.Refresh()
                        time.sleep(0.5)
                        workbook.Close(SaveChanges=True)
                        print("    图表数据更新成功")

                    except Exception as e:
                        print(f"    图表错误：{str(e)}")
                        try:
                            workbook.Close(SaveChanges=False)
                        except:
                            pass

        # 保存并关闭
        time.sleep(1)
        presentation.Save()
        presentation.Close()
        ppt_app.Quit()
        print(f"\nPPT已保存至: {ppt_path}")

    except Exception as e:
        print(f"错误：{e}")
        try:
            presentation.Close()
            ppt_app.Quit()
        except:
            pass


if __name__ == "__main__":
    ppt_path = r"D:\工作记录\【RPA】亚马逊竞对数据抓取\新-2025\数据分析报告模板.pptx"
    excel_path = r"D:\工作记录\【RPA】亚马逊竞对数据抓取\新-2025\数据分析报告底表.xlsx"
    update_ppt_from_excel(ppt_path, excel_path)