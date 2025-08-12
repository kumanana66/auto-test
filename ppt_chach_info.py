import win32com.client
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

def extract_ppt_data(ppt_path, output_excel_path):
    """从PPT中提取所有文本框、图表和表格数据，并保存到Excel的同一个工作表中"""
    try:
        if not os.path.exists(ppt_path):
            print(f"错误：PPT文件 '{ppt_path}' 不存在")
            return
        
        output_dir = os.path.dirname(output_excel_path)
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        ppt = win32com.client.Dispatch("PowerPoint.Application")
        ppt.Visible = True
        # ppt.DisplayAlerts = False
        presentation = ppt.Presentations.Open(ppt_path)
        print(f"已打开PPT: {ppt_path}")

        wb = Workbook()
        ws = wb.active
        ws.title = "汇总"

        headers = ["幻灯片编号", "类型", "形状名称", "内容/标题", "数据列1", "数据列2", "数据列3", "数据列4"]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        current_row = 2

        for slide_idx, slide in enumerate(presentation.Slides, 1):
            print(f"\n处理幻灯片 {slide_idx}...")
            
            for shape in slide.Shapes:
                shape_name = shape.Name
                print(f"  处理形状: {shape_name}")
                
                # 1. 处理表格
                if shape.HasTable:
                    try:
                        table = shape.Table
                        rows = table.Rows.Count
                        cols = table.Columns.Count
                        
                        print(f"    找到表格: {shape_name} ({rows}行{cols}列)")
                        
                        # 写入表格数据
                        for row in range(1, rows + 1):
                            ws.cell(row=current_row, column=1, value=slide_idx)
                            ws.cell(row=current_row, column=2, value="表格")
                            ws.cell(row=current_row, column=3, value=shape_name)
                            
                            # 如果是第一行，作为标题
                            if row == 1:
                                ws.cell(row=current_row, column=4, value="表头")
                            else:
                                ws.cell(row=current_row, column=4, value=f"数据行{row-1}")
                            
                            # 写入数据列
                            for col in range(1, min(cols, 8-4+1) + 1):
                                cell_text = table.Cell(row, col).Shape.TextFrame.TextRange.Text
                                ws.cell(row=current_row, column=4+col, value=cell_text)
                            
                            current_row += 1
                        
                    except Exception as e:
                        print(f"    提取表格数据失败: {e}")
                        ws.cell(row=current_row, column=1, value=slide_idx)
                        ws.cell(row=current_row, column=2, value="表格（错误）")
                        ws.cell(row=current_row, column=3, value=shape_name)
                        ws.cell(row=current_row, column=4, value=str(e))
                        current_row += 1
                
                # 2. 处理图表
                elif shape.HasChart:
                    try:
                        chart = shape.Chart
                        chart_type = get_chart_type_name(chart.ChartType)
                        print(f"    找到{chart_type}: {shape_name}")
                        
                        # 获取图表标题
                        chart_title = ""
                        if chart.HasTitle:
                            chart_title = chart.ChartTitle.Text
                        
                        # 激活图表数据
                        chart_data = chart.ChartData
                        chart_data.Activate()
                        
                        # 获取链接的Excel工作簿
                        workbook = chart_data.Workbook
                        worksheet = workbook.Worksheets(1)
                        
                        # 获取完整数据范围
                        data_range = worksheet.UsedRange
                        rows = data_range.Rows.Count
                        cols = data_range.Columns.Count
                        
                        # 写入图表标题行
                        ws.cell(row=current_row, column=1, value=slide_idx)
                        ws.cell(row=current_row, column=2, value=f"{chart_type}标题")
                        ws.cell(row=current_row, column=3, value=shape_name)
                        ws.cell(row=current_row, column=4, value=chart_title)
                        current_row += 1
                        
                        # 写入图表数据
                        for row in range(1, rows + 1):
                            ws.cell(row=current_row, column=1, value=slide_idx)
                            ws.cell(row=current_row, column=2, value=chart_type)
                            ws.cell(row=current_row, column=3, value=shape_name)
                            
                            # 如果是第一行，作为内容/标题
                            if row == 1:
                                ws.cell(row=current_row, column=4, value="表头")
                            else:
                                ws.cell(row=current_row, column=4, value=f"数据行{row-1}")
                            
                            # 写入数据列
                            for col in range(1, min(cols, 8-4+1) + 1):
                                cell_value = worksheet.Cells(row, col).Value
                                ws.cell(row=current_row, column=4+col, value=cell_value)
                            
                            current_row += 1

                        workbook.Close(SaveChanges=False)
                        
                    except Exception as e:
                        print(f"    提取图表数据失败: {e}")
                        ws.cell(row=current_row, column=1, value=slide_idx)
                        ws.cell(row=current_row, column=2, value=f"{chart_type}（错误）")
                        ws.cell(row=current_row, column=3, value=shape_name)
                        ws.cell(row=current_row, column=4, value=str(e))
                        current_row += 1
                
                # 3. 处理文本框
                elif shape.HasTextFrame:
                    text_frame = shape.TextFrame
                    if text_frame.HasText:
                        text = text_frame.TextRange.Text
                        ws.cell(row=current_row, column=1, value=slide_idx)
                        ws.cell(row=current_row, column=2, value="文本框")
                        ws.cell(row=current_row, column=3, value=shape_name)
                        ws.cell(row=current_row, column=4, value=text)
                        current_row += 1
                        print(f"    找到文本框: {text[:50]}...")
        
        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = min(adjusted_width, 50)
        
        wb.save(output_excel_path)
        print(f"\n数据已成功保存到: {output_excel_path}")

        presentation.Close()
        ppt.Quit()
        
    except Exception as e:
        print(f"发生错误: {e}")
        try:
            presentation.Close()
            ppt.Quit()
        except:
            pass

def get_chart_type_name(chart_type):
    """将图表类型ID转换为可读名称"""
    chart_types = {
        5: "饼图",
        3: "柱形图",
        4: "条形图",
        6: "折线图",
        8: "面积图",
        10: "散点图",
        11: "股价图",
        12: "曲面图",
        13: "圆环图",
        14: "雷达图",
        15: "气泡图",
        16: "股价图",
        18: "圆柱图",
        19: "圆锥图",
        20: "棱锥图",
        51: "柱形图"

    }
    return chart_types.get(chart_type, f"未知类型({chart_type})")

if __name__ == "__main__":
    # PPT文件路径
    ppt_path = r"D:\工作记录\【RPA】亚马逊竞对数据抓取\新-2025\数据分析报告模板.pptx"
    
    # 输出Excel文件路径
    output_excel_path = r"D:\工作记录\【RPA】亚马逊竞对数据抓取\新-2025\数据分析报告底表.xlsx"
    
    # 执行提取
    extract_ppt_data(ppt_path, output_excel_path)
